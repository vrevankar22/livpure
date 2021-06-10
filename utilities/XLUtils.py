import openpyxl
from utilities.readProperties import ReadConfig

# excel = ReadConfig.getExcelsheet()
excel = ".\\testdata\getData.xlsx"

def getRowCount(file,sheetName):
    workbook = openpyxl.load_workbook(excel)
    sheet = workbook[sheetName]
    return (sheet.max_row)

def getColumnCount(file,sheetName):
    workbook = openpyxl.load_workbook(excel)
    sheet = workbook[sheetName]
    return (sheet.max_column)

def readData(file,sheetName,rownum,columnnum):
    workbook = openpyxl.load_workbook(excel)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum,column=columnnum).value

def writeData(file,sheetName,rownum,columnnum,data):
    workbook = openpyxl.load_workbook(excel)
    sheet = workbook[sheetName]
    sheet.cell(row=rownum,column=columnnum).value = data
    workbook.save(file)
    workbook.close()